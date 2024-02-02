import pandas as pd
from sklearn.model_selection import StratifiedKFold
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, roc_auc_score, f1_score, precision_score, recall_score, confusion_matrix, matthews_corrcoef
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import LabelEncoder
from sklearn.svm import SVC
from sklearn.naive_bayes import GaussianNB
from sklearn.ensemble import AdaBoostClassifier
from sklearn.ensemble import BaggingClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import GridSearchCV
from imblearn.over_sampling import SMOTE
import matplotlib.pyplot as plt
from sklearn.metrics import roc_curve, auc
from sklearn.preprocessing import label_binarize
from scipy.stats import f_oneway
import seaborn as sns
from collections import OrderedDict
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from joblib import dump, load
from datetime import datetime
import numpy as np
import os

# Classifiers to be evaluated
classifiers = {
    "GradientBoosting": GradientBoostingClassifier(),
    "LogisticRegression": LogisticRegression(),
    "SVM": SVC(),
    "KNN": KNeighborsClassifier(),
    "DecisionTree": DecisionTreeClassifier(),
    "GaussianNaiveBayes": GaussianNB(),
    "AdaBoost": AdaBoostClassifier(),
    "Bagging": BaggingClassifier(),
    "MLPClassifier": MLPClassifier(),
    "RandomForestClassifier": RandomForestClassifier(),
}

param_grids = {
    "GradientBoosting": {
        "n_estimators": [700, 800],
        "learning_rate": [0.2, 0.3],
        "max_depth": [20, 25],
        "min_samples_split": [2, 4],
        "min_samples_leaf": [1, 2]
    },
    "LogisticRegression": {
        "C": [300, 400],
        "max_iter": [700, 800],
        "solver": ["newton-cg", "lbfgs"],
        "penalty": ["l2"]
    },
    "SVM": {
        "C": [300, 400],
        "kernel": ["rbf", "poly"],
        "gamma": [0.01, 0.1],
        "probability": [True],
        "class_weight": ['balanced']
    },
    "KNN": {
        "n_neighbors": [25, 30],
        "weights": ["distance"],
        "p": [1, 2]
    },
    "DecisionTree": {
        "max_depth": [25, 30],
        "criterion": ["gini", "entropy"],
        "min_samples_split": [2, 4],
        "min_samples_leaf": [1, 2]
    },
    "AdaBoost": {
        "n_estimators": [700, 800],
        "learning_rate": [0.2, 0.3],
        "algorithm": ["SAMME.R"]
    },
    "Bagging": {
        "n_estimators": [700, 800],
        "max_features": [0.5, 1.0],
        "max_samples": [0.5, 1.0],
        "bootstrap": [True]
    },
    "MLPClassifier": {
        "solver": ["sgd", "adam"],
        "alpha": [0.01, 0.1],
        "max_iter": [500, 600],
        "hidden_layer_sizes": [(400,), (400, 400)],
        "learning_rate_init": [0.01],
        "early_stopping": [True],
        "batch_size": [64, 128]
    },
    "RandomForestClassifier": {
        "n_estimators": [700, 800],
        "max_depth": [20, 25],
        "min_samples_split": [2, 4],
        "min_samples_leaf": [1, 2],
        "criterion": ["gini", "entropy"]
    }
}

# Evaluation metrics
metrics = OrderedDict([
    ("AUC", roc_auc_score),
    ("CA", accuracy_score),
    ("F1", f1_score),
    ("Precision", precision_score),
    ("Recall", recall_score),
    ("MCC", matthews_corrcoef),
])

# Define color fills
header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Light orange color for headers
data_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")    # Light blue color for data

# Define font and alignment
header_font = Font(bold=True)
left_aligned_text = Alignment(horizontal="left")

# Define a thin border
thin_border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

# Function to apply borders
def apply_borders(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = thin_border

# Adjusts the column widths in the given worksheet such that each column is equal to MAX_CONTENT_LENGTH + BUFFER
def adjust_column_widths(worksheet, buffer=5):
    max_length = 0
    for column_cells in worksheet.columns:
        for cell in column_cells:
            try:
                length = len(str(cell.value))
                max_length = max(length, max_length)
            except:
                pass

    adjusted_width = max_length + buffer

    for column_cells in worksheet.columns:
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = adjusted_width

# Color the "Model Performances" sheet
def color_model_performances_sheet(worksheet):
    # Apply color fill to headers
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = left_aligned_text

    # Apply color fill to classifier names and scores
    for row in worksheet.iter_rows(min_row=2, max_col=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = left_aligned_text

    for row in worksheet.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.fill = data_fill
            cell.alignment = left_aligned_text

# Color the "Confusion Matrix" sheet
def color_confusion_matrix_sheet(worksheet):
    # Apply styles to the specified cells
    # All cells with strings "Best Classifier", "Predicted / Actual", "A", "B", "C", "D"
    header_cells = ['A1', 'A3', 'A4', 'A5', 'A6', 'A7', 'B3', 'C3', 'D3', 'E3']
    for cell in header_cells:
        worksheet[cell].fill = header_fill
        worksheet[cell].font = header_font
        worksheet[cell].alignment = left_aligned_text

    # Apply styles for best classifier value cell
    worksheet["B1"].fill = data_fill
    worksheet["B1"].alignment = left_aligned_text

    # Apply styles to numeric data cells
    for row in worksheet.iter_rows(min_row=4, min_col=2, max_row=7, max_col=5):
        for cell in row:
            cell.fill = data_fill
            cell.alignment = left_aligned_text

def output_performance_excel(test_set_name, X_test, y_test, results_dict, filename):
    # Print reserved test set scores
        print(f"{test_set_name} scores:\n")
        for clf_name, clf_results in results_dict.items():
            print(f"Results for {clf_name}:")
            for metric_name, score in clf_results.items():
                print(f"\t{metric_name}: {score}")

        # Determine the best classifier based on Classification Accuracy
        best_clf_name = max(results_dict, key=lambda x: results_dict[x]['CA'])
        best_clf = classifiers[best_clf_name]

        # Generate the confusion matrix for the best classifier
        y_pred = best_clf.predict(X_test)
        conf_matrix = confusion_matrix(y_test, y_pred)

        # Save results to Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Model Performance"

        # Write headers for model performance
        metric_names = list(metrics.keys())
        ws.append(['Classifier'] + metric_names)  # Adding column headers

        # Write model performance to the worksheet
        for clf_name, clf_results in results_dict.items():
            row = [clf_name] + [clf_results[metric] for metric in metrics]
            ws.append(row)

        # Assuming 'le' is your LabelEncoder instance
        class_name_mapping = {index: label for index, label in enumerate(le.classes_)}

        # Add confusion matrix in a new sheet
        ws_conf = wb.create_sheet("Confusion Matrix")

        # Specify the best classifier
        ws_conf.append(["Best Classifier:", best_clf_name])

        # Empty row for readability
        ws_conf.append([])

        # Add headers for confusion matrix with class names
        ws_conf.append(["Predicted / Actual"] + [class_name_mapping[i] for i in range(conf_matrix.shape[0])])

        # Add confusion matrix data with class names
        for i, row in enumerate(conf_matrix):
            ws_conf.append([class_name_mapping[i]] + list(row))

        # Finally adjust column width before saving
        adjust_column_widths(ws)
        adjust_column_widths(ws_conf)

        # Color code sheets
        color_model_performances_sheet(ws)
        color_confusion_matrix_sheet(ws_conf)

        # Apply borders to show grid view
        apply_borders(ws, f"A1:G{1 + len(results_dict)}")
        apply_borders(ws_conf, "A1:B1")
        apply_borders(ws_conf, "A3:E7")

        # Save workbook with the timestamped filename
        wb.save(f"./output/{filename}")

        print(f"According to {test_set_name}:\n")

        # Print out the best classifier and its confusion matrix
        print(f"Best classifier: {best_clf_name}")
        print("Confusion Matrix for the best classifier:")
        print(conf_matrix)

# Script starts here
if __name__ == "__main__":
    # Load data
    data = pd.read_excel("EVSurveyData.xlsx")

    # Replace "?" with NaN
    data.replace("?", np.nan, inplace=True)

    # Impute missing values
    imputer = SimpleImputer(strategy='most_frequent')
    data_imputed = pd.DataFrame(imputer.fit_transform(data), columns=data.columns)

    # Encode target variable
    le = LabelEncoder()
    target = le.fit_transform(data_imputed['Q16'])

    # Drop target column from features
    features = data_imputed.drop('Q16', axis=1)

    # List of ordinal columns that shouldn't go through encoding
    ordinal_columns = ["Q14", "Q15", "Q17", "Q18_1", "Q18_2", "Q18_3", "Q18_4", "Q18_5", "Q18_6", "Q18_7", "Q18_8",
                    "Q18_9", "Q18_10", "Q18_11", "Q18_12", "Q18_13", "Q18_14", "Q18_15", "Q18_16", "Q18_17", 
                    "Q18_18", "Q18_19", "Q18_20", "Q18_21", "Q18_22", "Q18_23", "Q20", "Q21"]

    # List of nominal categorical columns that need one-hot encoding
    nominal_columns = [col for col in features.columns if col not in ordinal_columns]

    # Apply one-hot encoding to nominal columns only
    features_encoded = pd.get_dummies(features[nominal_columns])

    # Combine the one-hot encoded columns with the ordinal columns
    features_encoded = pd.concat([features[ordinal_columns].copy(), features_encoded], axis=1)

    # Convert the floats in the final DataFrame to int decrease computational cost
    for col in ordinal_columns:
        features_encoded[col] = features_encoded[col].astype(int)

    # Apply SMOTE to data
    smote = SMOTE(random_state=42)
    features_encoded, target = smote.fit_resample(features_encoded, target)

    # Stratified K-Fold cross-validation
    kfold = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)

    # Dictionary of models that weren't found as a .joblib file
    models_to_retrain = {}

    # Init the objects
    X_split_test = features_encoded
    y_split_test = target
    features_encoded = features_encoded
    target = target

    # Important features array
    top_features = []
    split_indices = []
    important_features_path = f"./models/MostImportantFeatures.joblib"
    train_test_split_indices = f"./models/ReservedSampleIndices.joblib"

    # If the most important features are known, load them. If not, feature
    # selection is in order
    if os.path.exists(important_features_path) and os.path.exists(train_test_split_indices):
        top_features = load(important_features_path)
        features_encoded = features_encoded[top_features]

        split_indices = load(train_test_split_indices)
        X_split_test = features_encoded.iloc[split_indices[1]]
        y_split_test = target[split_indices[1]]

        features_encoded = features_encoded.iloc[split_indices[0]]
        target = target[split_indices[0]]
    else:
        # Feature selection
        # Train a RandomForestClassifier
        print("Performing feature selection... ", end="")
        rf = RandomForestClassifier(n_estimators=100, random_state=42)
        rf.fit(features_encoded, target)

        # Extract feature importances
        feature_importances = rf.feature_importances_

        # Select the top 50 features
        indices = np.argsort(feature_importances)[::-1][:50]
        top_features = features_encoded.columns[indices]

        # Plot the importances of the top 50 features
        plt.figure(figsize=(10, 15))
        sns.barplot(x=feature_importances[indices], y=top_features)
        plt.title('Top 50 Feature Importances')
        plt.xlabel('Relative Importance')
        plt.ylabel('Feature Name')
        plt.savefig("./assets/top_50_feature_importances.png")
        plt.close()

        # Save the most important features
        dump(top_features, important_features_path)

        # Reserve 200 samples to test the models on once they are loaded
        train_indices, test_indices = train_test_split(range(features_encoded.shape[0]), test_size=200, random_state=42, stratify=target)
        X_split_test = features_encoded.iloc[test_indices]
        y_split_test = target[test_indices]

        features_encoded = features_encoded.iloc[train_indices]
        target = target[train_indices]

        # Populate the array
        split_indices.append(train_indices)
        split_indices.append(test_indices)

        # Save the resulting indices
        dump(split_indices, train_test_split_indices)

        # Overwrite features_encoded to only contain the selected features
        features_encoded = features_encoded[top_features]
        print("Done\n")

        # Since the selected features may not always be the same for the loaded models, the models
        # need retraining, even the loaded ones
        models_to_retrain = classifiers

    # Results dictionaries
    # KFold results
    results = {}

    # Reserved test set scores
    results_reserved = {}

    output_reserved_validation_scores = False
    if len(models_to_retrain) == 0:
        # Try to find the models, if found, calculate their models and if not, queue them to be retrained
        for model_name, clf in classifiers.items():
            model_path = f"./models/{model_name}.joblib"
            if os.path.exists(model_path):
                output_reserved_validation_scores = True

                # Model file found, load it
                model = load(model_path)
                classifiers[model_name] = model

                # Initialize metrics container
                results_reserved[model_name] = {}
                
                for metric_name, metric in metrics.items():
                    scores = {metric: [] for metric in metrics}

                    y_pred = model.predict(X_split_test)

                    # Metric calculations
                    for metric_name, metric_func in metrics.items():
                        if metric_name == "AUC":
                            y_pred_proba = model.predict_proba(X_split_test)
                            score = metric_func(y_split_test, y_pred_proba, multi_class="ovo")
                        elif metric_name in ["CA", "MCC"]:
                            score = metric_func(y_split_test, y_pred)
                        else:
                            score = metric_func(y_split_test, y_pred, average="weighted")
                            
                        scores[metric_name].append(score)
                    
                    results_reserved[model_name] = {metric: np.mean(values) for metric, values in scores.items()}

                print(f"Loaded model: {model_name}")

            else:
                models_to_retrain[model_name] = clf

    output_kfold_cv_scores = False
    # For all models needing training...
    if len(models_to_retrain) > 0:
        print(f"\n{len(models_to_retrain)} models missing, training them ...\n")

        output_kfold_cv_scores = True

        # Run cross-validation for each classifier
        for clf_name, clf in models_to_retrain.items():
            print(f"Training {clf_name} ... ")
            
            # This is the instance of the classifier that will be filled after grid search
            best_clf = clf

            # Hyperparameter tuning
            if clf_name != "GaussianNaiveBayes":
                print("Running GridSearchCV for hyperparameter tuning")

                # Update parameter grid for grid search
                param_grid = param_grids[clf_name]

                # GridSearchCV for the model
                grid_search = GridSearchCV(clf, param_grid, scoring="accuracy", cv=kfold, n_jobs=-1)
                grid_search.fit(features_encoded, target)

                # Best estimator after grid search
                best_clf = grid_search.best_estimator_

                print(f"Finished tuning parameters for {clf_name}")

            # Fitting and scoring
            scores = {metric: [] for metric in metrics}

            for train_index, test_index in kfold.split(features_encoded, target):
                X_train, X_test = features_encoded.iloc[train_index], features_encoded.iloc[test_index]
                y_train, y_test = target[train_index], target[test_index]

                # Fit model
                best_clf.fit(X_train, y_train)
                y_pred = best_clf.predict(X_test)

                # Metric calculations
                for metric_name, metric_func in metrics.items():
                    if metric_name == "AUC":
                        y_pred_proba = best_clf.predict_proba(X_test)
                        score = metric_func(y_test, y_pred_proba, multi_class="ovo")
                    elif metric_name in ["CA", "MCC"]:
                        score = metric_func(y_test, y_pred)
                    else:
                        score = metric_func(y_test, y_pred, average="weighted")
                    scores[metric_name].append(score)

            results[clf_name] = {metric: np.mean(values) for metric, values in scores.items()}

            
            # Save model
            classifiers[clf_name] = best_clf
            dump(best_clf, f"./models/{clf_name}.joblib")
        
            print(f"Training done for {clf_name}\n")
        print("\n")

    # Get the current date and time
    current_time = datetime.now()

    # Format the date and time in a human-readable format (e.g., YYYYMMDD_HHMMSS)
    time_str = current_time.strftime("%Y%m%d_%H%M%S")

    # Print results
    if output_kfold_cv_scores:
        # Print KFold CV scores
        print("KFold cross-validation scores (averaged):\n")
        for clf_name, clf_results in results.items():
            print(f"Results for {clf_name}:")
            for metric_name, score in clf_results.items():
                print(f"\t{metric_name}: {score}")

        # Binarize the output
        y_bin = label_binarize(target, classes=np.unique(target))
        n_classes = y_bin.shape[1]

        # Get the actual classes
        class_labels = le.classes_

        # Initialize plot for ROC curves
        plt.figure(figsize=(10, 8))

        # Compute ROC curve and ROC area for each class
        lines = []  # List to hold the line objects
        labels = []  # List to hold the label objects

        # Calculate, save and print performance on X_split_test, which was not seen to this point
        print("\nPlotting ROC Curves... ", end="")
        for clf_name, clf in classifiers.items():
            # Check if classifier supports the prediction of probabilities
            if hasattr(clf, "predict_proba"):
                probas = clf.predict_proba(features_encoded)

                # Compute ROC curve and AUC for each class
                for i in range(n_classes):
                    fpr, tpr, _ = roc_curve(y_bin[:, i], probas[:, i])
                    roc_auc = auc(fpr, tpr)
                    line, = plt.plot(fpr, tpr, lw=2, label=f'ROC curve of {class_labels[i]} for {clf_name} (AUC = {roc_auc:.2f})')
                    lines.append(line)
                    labels.append(f'ROC curve of {class_labels[i]} for {clf_name} (AUC = {roc_auc:.2f})')

        # Plot random chance line
        plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')

        # Add labels and title
        plt.xlabel('False Positive Rate')
        plt.ylabel('True Positive Rate')
        plt.title('Receiver Operating Characteristic (ROC) Curves for Multi-Class')

        # Save the plot to a file without the legend
        plt.savefig("./assets/roc_curves_multiclass.png", bbox_inches='tight')

        # Close the plot
        plt.close()

        # Create a new figure for the legend
        plt.figure(figsize=(8, 4))
        plt.figlegend(lines, labels, loc='center', ncol=1, labelspacing=0.)

        # Save the legend to a file
        plt.savefig("./assets/roc_legend.png", bbox_inches='tight')

        # Close the plot
        plt.close()
        print("Done\n")

        print("Plotting Target and Important Feature Distributions... ", end="")

        # Get actual target vector
        class_labels = le.inverse_transform(target)

        # Visualize the distribution of the target variable
        plt.figure(figsize=(8, 6))
        sns.countplot(x=class_labels)
        plt.title("Distribution of Target Variable")
        plt.xlabel("Classes")
        plt.ylabel("Count")
        plt.savefig("./assets/target_distribution.png")
        plt.close()

        # Visualize the distribution of key features
        top_features_to_visualize = top_features[:15]

        for feature in top_features_to_visualize:
            plt.figure(figsize=(8, 6))
            sns.histplot(data=features_encoded, x=feature, kde=True)
            plt.title(f"Distribution of Feature: {feature}")
            plt.xlabel(feature)
            plt.ylabel("Count")
            plt.savefig(f"./assets/distribution_{feature}.png")
            plt.close()

        # Visualize correlations between features
        plt.figure(figsize=(12, 10))

        # Concatenate target with top features for correlation analysis
        combined_data = pd.concat([features_encoded[top_features_to_visualize], pd.Series(target, name='Target (Q16)')], axis=1)

        # Compute the correlation matrix including the target variable
        correlation_matrix = combined_data.corr()

        # Generate the heatmap
        sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm')
        plt.title("Correlation Heatmap of Top Features with Target")
        plt.savefig("./assets/correlation_heatmap_with_target.png")
        plt.close()

        # Convert target to pandas Series
        target_series = pd.Series(target, name='Target')
        

        # Reset indices to ensure alignment, if necessary.
        features_encoded.reset_index(drop=True, inplace=True)
        target_series.reset_index(drop=True, inplace=True)

        # Initialize a dictionary to store ANOVA results for the top 15 features.
        anova_results_top_features = {}

        # Perform ANOVA for each of the top 15 features.
        for feature in top_features_to_visualize:
            # Create a list of arrays, each containing the values of the feature for a specific class.
            groups = [features_encoded.loc[target_series == cls, feature].values for cls in np.unique(target_series)]

            # Perform the ANOVA test and store the p-value.
            f_statistic, p_value = f_oneway(*groups)
            anova_results_top_features[feature] = p_value

        # Convert p-values to -log(p-value) for visualization.
        # Handle the case where p-value is 0 by setting a minimum p-value.
        min_p_value = np.nextafter(0, 1)  # Smallest positive float number > 0
        anova_df_top = pd.DataFrame.from_dict(anova_results_top_features, orient='index', columns=['p_value'])
        anova_df_top['p_value'] = anova_df_top['p_value'].replace(0, min_p_value)
        anova_df_top['-log(p_value)'] = -np.log10(anova_df_top['p_value'])

        # Sort the results for better visualization.
        anova_df_top_sorted = anova_df_top.sort_values('-log(p_value)', ascending=False)

        # Plotting.
        plt.figure(figsize=(12, 6))
        anova_df_top_sorted['-log(p_value)'].plot(kind='bar')
        plt.title('ANOVA Test Results (-log(p-value)) for Top 15 Features')
        plt.ylabel('-log(p-value)')
        plt.xlabel('Top 15 Features')
        plt.xticks(rotation=45)
        plt.tight_layout()  # This ensures the labels are not cut off.
        plt.savefig("./assets/anova_test_results.png")
        plt.close()

        # Concatenate the time string with the filename
        filename = f"classification_results_{time_str}.xlsx"

        output_performance_excel("KFold Cross-Validation", features_encoded, target, results, filename)

    if output_reserved_validation_scores:      
        # Concatenate the time string with the filename
        filename = f"reserved_classification_results_{time_str}.xlsx"

        output_performance_excel("Reserved Test Set", X_split_test, y_split_test, results_reserved, filename)
        